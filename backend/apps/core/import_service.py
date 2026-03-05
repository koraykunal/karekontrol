import uuid
from datetime import datetime

from django.db import transaction
from openpyxl import load_workbook

from apps.entities.models import Entity
from apps.entities.services import EntityService
from apps.organizations.models import Department
from apps.procedures.models import Procedure, ProcedureStep
from apps.core.constants import ProcedurePriority, IntervalUnit


ENTITY_TYPES = {"equipment", "vehicle", "facility", "tool", "device", "other"}
ENTITY_STATUSES = {"active", "inactive", "maintenance", "retired"}


def _parse_date(value, field_name, row_num):
    """Parse a date value from Excel cell. Returns (date|None, error|None)."""
    if value is None or (isinstance(value, str) and value.strip() == ''):
        return None, None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, str):
        for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(value.strip(), fmt).date(), None
            except ValueError:
                continue
        return None, {
            'row': row_num,
            'field': field_name,
            'message': f'Geçersiz tarih formatı: "{value}". GG.AA.YYYY formatında olmalı',
        }
    return None, {
        'row': row_num,
        'field': field_name,
        'message': f'Geçersiz tarih değeri',
    }


def _parse_bool(value):
    """Parse a boolean from 'Evet'/'Hayır' or True/False."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in ('evet', 'true', '1', 'yes')


def _parse_int(value, field_name, row_num, required=False):
    """Parse integer. Returns (int|None, error|None)."""
    if value is None or (isinstance(value, str) and value.strip() == ''):
        if required:
            return None, {
                'row': row_num,
                'field': field_name,
                'message': 'Bu alan zorunludur',
            }
        return None, None
    try:
        v = int(float(value))
        return v, None
    except (ValueError, TypeError):
        return None, {
            'row': row_num,
            'field': field_name,
            'message': f'Geçerli bir sayı olmalı',
        }


def _cell_str(value):
    """Get trimmed string from cell value, or empty string if None."""
    if value is None:
        return ''
    return str(value).strip()


def _validate_headers(ws, expected, sheet_name):
    """Validate that the first row contains expected headers. Returns list of errors."""
    actual = [_cell_str(ws.cell(row=1, column=c).value).lower() for c in range(1, len(expected) + 1)]
    expected_lower = [h.lower() for h in expected]
    missing = [h for h, a in zip(expected, actual) if a != h.lower()]
    if missing:
        return [{
            'row': 1,
            'field': 'headers',
            'message': f'"{sheet_name}" sheet\'inde eksik veya yanlış sırada kolonlar: {", ".join(missing)}',
        }]
    return []


class EntityImportService:
    REQUIRED_HEADERS = [
        "name", "code", "entity_type", "department_name", "status",
        "location", "serial_number", "manufacturer", "model",
        "purchase_date", "warranty_expiry_date", "description", "notes",
    ]

    @staticmethod
    def process(file, organization, user):
        try:
            wb = load_workbook(file, data_only=True)
        except Exception:
            return {
                'success': False,
                'message': 'Dosya okunamadı. Geçerli bir .xlsx dosyası yükleyin',
                'data': {'errors': []},
            }

        if 'Varlıklar' not in wb.sheetnames:
            return {
                'success': False,
                'message': '"Varlıklar" adlı sheet bulunamadı',
                'data': {'errors': []},
            }

        ws = wb['Varlıklar']
        errors = _validate_headers(ws, EntityImportService.REQUIRED_HEADERS, 'Varlıklar')
        if errors:
            return {'success': False, 'message': 'Kolon başlıkları hatalı', 'data': {'errors': errors}}

        # Pre-fetch departments for this org
        departments = {d.name.lower(): d for d in Department.objects.filter(organization=organization)}

        # Pre-fetch existing codes
        existing_codes = set(
            Entity.objects.filter(organization=organization, is_deleted=False)
            .values_list('code', flat=True)
        )

        rows = []
        seen_codes = set()
        data_start_row = 4  # 1=header, 2=example, 3=hints

        for row_num in range(data_start_row, ws.max_row + 1):
            # Skip empty rows
            name = _cell_str(ws.cell(row=row_num, column=1).value)
            if not name:
                continue

            code = _cell_str(ws.cell(row=row_num, column=2).value)
            entity_type = _cell_str(ws.cell(row=row_num, column=3).value).upper()
            department_name = _cell_str(ws.cell(row=row_num, column=4).value)
            status = _cell_str(ws.cell(row=row_num, column=5).value).upper() or 'ACTIVE'
            location = _cell_str(ws.cell(row=row_num, column=6).value)
            serial_number = _cell_str(ws.cell(row=row_num, column=7).value)
            manufacturer = _cell_str(ws.cell(row=row_num, column=8).value)
            model = _cell_str(ws.cell(row=row_num, column=9).value)
            purchase_raw = ws.cell(row=row_num, column=10).value
            warranty_raw = ws.cell(row=row_num, column=11).value
            description = _cell_str(ws.cell(row=row_num, column=12).value)
            notes = _cell_str(ws.cell(row=row_num, column=13).value)

            row_errors = []

            # Required fields
            if not name:
                row_errors.append({'row': row_num, 'field': 'name', 'message': 'Ad zorunludur'})

            if not entity_type:
                row_errors.append({'row': row_num, 'field': 'entity_type', 'message': 'Varlık tipi zorunludur'})
            elif entity_type not in ENTITY_TYPES:
                row_errors.append({
                    'row': row_num, 'field': 'entity_type',
                    'message': f'Geçersiz varlık tipi: "{entity_type}". Geçerli: {", ".join(sorted(ENTITY_TYPES))}',
                })

            if not department_name:
                row_errors.append({'row': row_num, 'field': 'department_name', 'message': 'Departman adı zorunludur'})
            else:
                dept = departments.get(department_name.lower())
                if not dept:
                    row_errors.append({
                        'row': row_num, 'field': 'department_name',
                        'message': f'Departman bulunamadı: "{department_name}"',
                    })

            if status not in ENTITY_STATUSES:
                row_errors.append({
                    'row': row_num, 'field': 'status',
                    'message': f'Geçersiz durum: "{status}"',
                })

            # Auto-generate code if blank
            if not code:
                code = f"ENT-{uuid.uuid4().hex[:8].upper()}"

            if code.lower() in seen_codes:
                row_errors.append({
                    'row': row_num, 'field': 'code',
                    'message': f'Bu kod dosya içinde tekrar ediyor: "{code}"',
                })
            elif code in existing_codes:
                row_errors.append({
                    'row': row_num, 'field': 'code',
                    'message': f'Bu kod zaten kullanılıyor: "{code}"',
                })

            purchase_date, err = _parse_date(purchase_raw, 'purchase_date', row_num)
            if err:
                row_errors.append(err)

            warranty_date, err = _parse_date(warranty_raw, 'warranty_expiry_date', row_num)
            if err:
                row_errors.append(err)

            errors.extend(row_errors)
            seen_codes.add(code.lower())

            rows.append({
                'name': name,
                'code': code,
                'entity_type': entity_type,
                'department': departments.get(department_name.lower()) if department_name else None,
                'status': status,
                'location': location or None,
                'serial_number': serial_number or None,
                'manufacturer': manufacturer or None,
                'model': model or None,
                'purchase_date': purchase_date,
                'warranty_expiry_date': warranty_date,
                'description': description or None,
                'notes': notes or None,
            })

        if not rows:
            return {
                'success': False,
                'message': 'Dosyada veri satırı bulunamadı (4. satırdan itibaren)',
                'data': {'errors': []},
            }

        # Quota check
        current_count, qr_quota = EntityService.check_quota(organization)
        if qr_quota > 0 and (current_count + len(rows)) > qr_quota:
            remaining = max(0, qr_quota - current_count)
            errors.append({
                'row': 0,
                'field': 'quota',
                'message': f'QR kota aşımı. Kota: {qr_quota}, mevcut: {current_count}, '
                           f'eklenecek: {len(rows)}, kalan: {remaining}',
            })

        if errors:
            return {
                'success': False,
                'message': f'{len(errors)} hata bulundu, hiçbir kayıt oluşturulmadı',
                'data': {'errors': errors},
            }

        # All valid — create in a transaction
        with transaction.atomic():
            for row_data in rows:
                Entity.objects.create(organization=organization, **row_data)

        return {
            'success': True,
            'message': f'{len(rows)} varlık oluşturuldu',
            'data': {'created_count': len(rows)},
        }


class ProcedureImportService:
    PROC_HEADERS = [
        "entity_code", "title", "description", "priority",
        "interval_value", "interval_unit", "estimated_duration_minutes",
        "requires_approval", "tags",
    ]
    STEP_HEADERS = [
        "procedure_title", "step_order", "title", "description",
        "requires_photo", "requires_notes", "expected_duration_minutes",
    ]

    @staticmethod
    def process(file, organization, user):
        try:
            wb = load_workbook(file, data_only=True)
        except Exception:
            return {
                'success': False,
                'message': 'Dosya okunamadı. Geçerli bir .xlsx dosyası yükleyin',
                'data': {'errors': []},
            }

        if 'Prosedürler' not in wb.sheetnames:
            return {
                'success': False,
                'message': '"Prosedürler" adlı sheet bulunamadı',
                'data': {'errors': []},
            }
        if 'Adımlar' not in wb.sheetnames:
            return {
                'success': False,
                'message': '"Adımlar" adlı sheet bulunamadı',
                'data': {'errors': []},
            }

        ws_proc = wb['Prosedürler']
        ws_step = wb['Adımlar']

        errors = []
        errors.extend(_validate_headers(ws_proc, ProcedureImportService.PROC_HEADERS, 'Prosedürler'))
        errors.extend(_validate_headers(ws_step, ProcedureImportService.STEP_HEADERS, 'Adımlar'))
        if errors:
            return {'success': False, 'message': 'Kolon başlıkları hatalı', 'data': {'errors': errors}}

        # Pre-fetch entities
        entities_by_code = {
            e.code.lower(): e
            for e in Entity.objects.filter(organization=organization, is_deleted=False)
        }

        valid_priorities = {p.value for p in ProcedurePriority}
        valid_units = {u.value for u in IntervalUnit}

        # --- Parse procedures ---
        proc_rows = []
        proc_titles = {}  # title -> index in proc_rows for step matching
        data_start = 4

        for row_num in range(data_start, ws_proc.max_row + 1):
            entity_code = _cell_str(ws_proc.cell(row=row_num, column=1).value)
            title = _cell_str(ws_proc.cell(row=row_num, column=2).value)
            if not entity_code and not title:
                continue

            description = _cell_str(ws_proc.cell(row=row_num, column=3).value)
            priority = _cell_str(ws_proc.cell(row=row_num, column=4).value).upper()
            interval_raw = ws_proc.cell(row=row_num, column=5).value
            interval_unit = _cell_str(ws_proc.cell(row=row_num, column=6).value).upper()
            duration_raw = ws_proc.cell(row=row_num, column=7).value
            requires_approval = _parse_bool(ws_proc.cell(row=row_num, column=8).value)
            tags_raw = _cell_str(ws_proc.cell(row=row_num, column=9).value)

            row_errors = []

            if not entity_code:
                row_errors.append({'row': row_num, 'field': 'entity_code', 'message': 'Varlık kodu zorunludur'})
            elif entity_code.lower() not in entities_by_code:
                row_errors.append({
                    'row': row_num, 'field': 'entity_code',
                    'message': f'Varlık bulunamadı: "{entity_code}"',
                })

            if not title:
                row_errors.append({'row': row_num, 'field': 'title', 'message': 'Başlık zorunludur'})

            if not priority:
                row_errors.append({'row': row_num, 'field': 'priority', 'message': 'Öncelik zorunludur'})
            elif priority not in valid_priorities:
                row_errors.append({
                    'row': row_num, 'field': 'priority',
                    'message': f'Geçersiz öncelik: "{priority}". Geçerli: {", ".join(sorted(valid_priorities))}',
                })

            interval_value, err = _parse_int(interval_raw, 'interval_value', row_num, required=True)
            if err:
                row_errors.append(err)
            elif interval_value is not None and interval_value <= 0:
                row_errors.append({
                    'row': row_num, 'field': 'interval_value',
                    'message': 'Tekrar sıklığı 0\'dan büyük olmalı',
                })

            if not interval_unit:
                row_errors.append({'row': row_num, 'field': 'interval_unit', 'message': 'Tekrar birimi zorunludur'})
            elif interval_unit not in valid_units:
                row_errors.append({
                    'row': row_num, 'field': 'interval_unit',
                    'message': f'Geçersiz birim: "{interval_unit}". Geçerli: {", ".join(sorted(valid_units))}',
                })

            duration, err = _parse_int(duration_raw, 'estimated_duration_minutes', row_num)
            if err:
                row_errors.append(err)

            tags = [t.strip() for t in tags_raw.split(',') if t.strip()] if tags_raw else []

            errors.extend(row_errors)

            proc_titles[title.lower()] = len(proc_rows)
            proc_rows.append({
                'entity': entities_by_code.get(entity_code.lower()) if entity_code else None,
                'title': title,
                'description': description or None,
                'priority': priority,
                'interval_value': interval_value,
                'interval_unit': interval_unit,
                'estimated_duration_minutes': duration,
                'requires_approval': requires_approval,
                'tags': tags,
                '_steps': [],
            })

        if not proc_rows:
            return {
                'success': False,
                'message': 'Prosedürler sheet\'inde veri satırı bulunamadı',
                'data': {'errors': []},
            }

        # --- Parse steps ---
        total_steps = 0
        for row_num in range(data_start, ws_step.max_row + 1):
            proc_title = _cell_str(ws_step.cell(row=row_num, column=1).value)
            step_order_raw = ws_step.cell(row=row_num, column=2).value
            title = _cell_str(ws_step.cell(row=row_num, column=3).value)
            if not proc_title and not title:
                continue

            description = _cell_str(ws_step.cell(row=row_num, column=4).value)
            requires_photo = _parse_bool(ws_step.cell(row=row_num, column=5).value)
            requires_notes = _parse_bool(ws_step.cell(row=row_num, column=6).value)
            duration_raw = ws_step.cell(row=row_num, column=7).value

            row_errors = []

            if not proc_title:
                row_errors.append({
                    'row': row_num, 'field': 'procedure_title',
                    'message': 'Prosedür başlığı zorunludur (Adımlar sheet)',
                })
            elif proc_title.lower() not in proc_titles:
                row_errors.append({
                    'row': row_num, 'field': 'procedure_title',
                    'message': f'Eşleşen prosedür bulunamadı: "{proc_title}" (Adımlar sheet)',
                })

            step_order, err = _parse_int(step_order_raw, 'step_order', row_num, required=True)
            if err:
                row_errors.append(err)
            elif step_order is not None and step_order <= 0:
                row_errors.append({
                    'row': row_num, 'field': 'step_order',
                    'message': 'Adım sırası 0\'dan büyük olmalı (Adımlar sheet)',
                })

            if not title:
                row_errors.append({
                    'row': row_num, 'field': 'title',
                    'message': 'Adım başlığı zorunludur (Adımlar sheet)',
                })

            duration, err = _parse_int(duration_raw, 'expected_duration_minutes', row_num)
            if err:
                row_errors.append(err)

            errors.extend(row_errors)

            if proc_title and proc_title.lower() in proc_titles:
                idx = proc_titles[proc_title.lower()]
                proc_rows[idx]['_steps'].append({
                    'step_order': step_order,
                    'title': title,
                    'description': description or None,
                    'requires_photo': requires_photo,
                    'requires_notes': requires_notes,
                    'expected_duration_minutes': duration,
                })
                total_steps += 1

        if errors:
            return {
                'success': False,
                'message': f'{len(errors)} hata bulundu, hiçbir kayıt oluşturulmadı',
                'data': {'errors': errors},
            }

        # All valid — create in a transaction
        with transaction.atomic():
            for proc_data in proc_rows:
                steps = proc_data.pop('_steps')
                entity = proc_data.pop('entity')
                procedure = Procedure.objects.create(
                    organization=organization,
                    entity=entity,
                    created_by=user,
                    **proc_data,
                )
                for step_data in steps:
                    ProcedureStep.objects.create(
                        procedure=procedure,
                        **step_data,
                    )

        return {
            'success': True,
            'message': f'{len(proc_rows)} prosedür ve {total_steps} adım oluşturuldu',
            'data': {
                'created_count': len(proc_rows),
                'steps_count': total_steps,
            },
        }
