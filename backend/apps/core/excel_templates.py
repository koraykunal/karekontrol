from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

EXAMPLE_FONT = Font(color="333333", size=10)
HINT_FONT = Font(italic=True, color="888888", size=9)
HINT_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _style_header_row(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_example_row(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER


def _style_hint_row(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = HINT_FONT
        cell.fill = HINT_FILL
        cell.alignment = Alignment(wrap_text=True)
        cell.border = THIN_BORDER


def _set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _add_dropdown(ws, col_letter, values, start_row=4, end_row=1000):
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=True,
    )
    dv.error = "Geçersiz değer. Lütfen listeden seçin."
    dv.errorTitle = "Geçersiz Giriş"
    dv.prompt = "Listeden bir değer seçin"
    dv.promptTitle = "Seçim"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


ENTITY_TYPES = ["equipment", "vehicle", "facility", "tool", "device", "other"]
ENTITY_STATUSES = ["active", "inactive", "maintenance", "retired"]
PRIORITIES = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
INTERVAL_UNITS = ["DAYS", "WEEKS", "MONTHS", "YEARS"]
YES_NO = ["Evet", "Hayır"]


def generate_entity_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Varlıklar"

    headers = [
        "name", "code", "entity_type", "department_name", "status",
        "location", "serial_number", "manufacturer", "model",
        "purchase_date", "warranty_expiry_date", "description", "notes",
    ]
    example = [
        "CNC Torna Makinesi", "CNC-001", "equipment", "Üretim",
        "active", "A Blok, Kat 2", "SN-123456", "Siemens",
        "Sinumerik 840D", "15.01.2024", "15.01.2026",
        "5 eksenli CNC torna", "Yıllık bakım gerekli",
    ]
    hints = [
        "Zorunlu", "Boş bırakılırsa otomatik oluşturulur",
        "equipment, vehicle, facility, tool, device, other",
        "Zorunlu - organizasyondaki departman adı",
        "active (varsayılan), inactive, maintenance, retired",
        "İsteğe bağlı", "İsteğe bağlı", "İsteğe bağlı", "İsteğe bağlı",
        "GG.AA.YYYY formatında", "GG.AA.YYYY formatında",
        "İsteğe bağlı", "İsteğe bağlı — BU SATIRI SİLİN",
    ]
    widths = [25, 15, 18, 20, 14, 20, 18, 18, 18, 16, 20, 30, 30]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)
    for col, val in enumerate(hints, 1):
        ws.cell(row=3, column=col, value=val)

    col_count = len(headers)
    _style_header_row(ws, col_count)
    _style_example_row(ws, col_count)
    _style_hint_row(ws, col_count)
    _set_column_widths(ws, widths)

    # Dropdowns
    _add_dropdown(ws, "C", ENTITY_TYPES)   # entity_type
    _add_dropdown(ws, "E", ENTITY_STATUSES)  # status

    ws.freeze_panes = "A2"

    return wb


def generate_procedure_template():
    wb = Workbook()

    # --- Sheet 1: Prosedürler ---
    ws1 = wb.active
    ws1.title = "Prosedürler"

    proc_headers = [
        "entity_code", "title", "description", "priority",
        "interval_value", "interval_unit", "estimated_duration_minutes",
        "requires_approval", "tags",
    ]
    proc_example = [
        "CNC-001", "Haftalık Bakım Kontrolü",
        "CNC torna makinesi haftalık bakım prosedürü",
        "HIGH", "7", "DAYS", "45", "Hayır",
        "bakım, cnc, haftalık",
    ]
    proc_hints = [
        "Zorunlu - varlık kodu", "Zorunlu", "İsteğe bağlı",
        "LOW, MEDIUM, HIGH, CRITICAL",
        "Zorunlu - pozitif sayı", "DAYS, WEEKS, MONTHS, YEARS",
        "Dakika cinsinden", "Evet / Hayır",
        "Virgülle ayrılmış — BU SATIRI SİLİN",
    ]
    proc_widths = [18, 30, 35, 14, 16, 16, 26, 18, 30]

    for col, header in enumerate(proc_headers, 1):
        ws1.cell(row=1, column=col, value=header)
    for col, val in enumerate(proc_example, 1):
        ws1.cell(row=2, column=col, value=val)
    for col, val in enumerate(proc_hints, 1):
        ws1.cell(row=3, column=col, value=val)

    col_count_1 = len(proc_headers)
    _style_header_row(ws1, col_count_1)
    _style_example_row(ws1, col_count_1)
    _style_hint_row(ws1, col_count_1)
    _set_column_widths(ws1, proc_widths)

    _add_dropdown(ws1, "D", PRIORITIES)       # priority
    _add_dropdown(ws1, "F", INTERVAL_UNITS)   # interval_unit
    _add_dropdown(ws1, "H", YES_NO)           # requires_approval

    ws1.freeze_panes = "A2"

    # --- Sheet 2: Adımlar ---
    ws2 = wb.create_sheet(title="Adımlar")

    step_headers = [
        "procedure_title", "step_order", "title", "description",
        "requires_photo", "requires_notes", "expected_duration_minutes",
    ]
    step_example = [
        "Haftalık Bakım Kontrolü", "1", "Yağ Seviyesi Kontrolü",
        "Yağ seviyesini kontrol edin ve gerekirse tamamlayın",
        "Evet", "Hayır", "10",
    ]
    step_hints = [
        "Prosedürler sheet'indeki title ile eşleşmeli", "Zorunlu - sıra numarası (1,2,3...)",
        "Zorunlu", "İsteğe bağlı", "Evet / Hayır",
        "Evet / Hayır", "Dakika — BU SATIRI SİLİN",
    ]
    step_widths = [30, 14, 25, 35, 16, 16, 26]

    for col, header in enumerate(step_headers, 1):
        ws2.cell(row=1, column=col, value=header)
    for col, val in enumerate(step_example, 1):
        ws2.cell(row=2, column=col, value=val)
    for col, val in enumerate(step_hints, 1):
        ws2.cell(row=3, column=col, value=val)

    col_count_2 = len(step_headers)
    _style_header_row(ws2, col_count_2)
    _style_example_row(ws2, col_count_2)
    _style_hint_row(ws2, col_count_2)
    _set_column_widths(ws2, step_widths)

    _add_dropdown(ws2, "E", YES_NO)  # requires_photo
    _add_dropdown(ws2, "F", YES_NO)  # requires_notes

    ws2.freeze_panes = "A2"

    return wb
